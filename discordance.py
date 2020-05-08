#!/usr/bin/env python
import csv
import sys
import os
from openpyxl import load_workbook, Workbook
import datetime
import errno
from time import sleep

discordance_master_list = []
sample_disease_dict = {}
failures_master_list = []


# Checks for consistency in tech called genotypes accross multiple tests

def main():
    print('This program checks .xlsx files, not .csv files. \nFilename must include "_Map.xlsx". Make sure calls are made in the right file')
    print('\nColumns must be as follows: sample in B, disease in C, genotype in H\n')
    sleep(4)
    maplist = getmaplist()
    try:
        for maps in maplist:
            gelmap = open_gel(maps)
            print("\nChecking {}".format((maps.split("\\")[-1])))
            sample_disease(gelmap)
    except PermissionError:
        print('Close the maps and try again.')
        sleep(4)
        quit()
    print("\nAll failures:")
    for item in failures_master_list:
        print(item)
    print("\nAll discordancies:")
    for item in discordance_master_list:
        print(item)
    if len(maplist) > 3:
        print("\nThere are more than 3 maps in the directory. All maps are referened for failure and discordance master lists. Consider deleting replicates or incomplete maps.")
    end = input("\nPress enter to end the program")


def getmaplist():
    # path to folder of genotype calls
    yesno = "x"
    while yesno.lower() not in "yesno":
        print("Attempt automatic folder detection?")
        yesno = input(":")
        if yesno.lower() not in "yesno":
            print("What?")
    maps = []
    if yesno.lower() in "yes":
        today = datetime.date.today()
        monday_date = today - datetime.timedelta(days=today.weekday())
        week_of = "Week of {}-{}-{}".format(monday_date.strftime("%m"), monday_date.strftime("%d"), monday_date.strftime("%y"))
        if today.weekday() != 0 and today.weekday() != 2:
            previous_date = today - datetime.timedelta(days=1)
        else:
            previous_date = today    
        previous_folder = "{}-{}-{}".format(previous_date.strftime("%m"), previous_date.strftime("%d"), previous_date.strftime("%y"))
        try:
            os.chdir('../../Current Year/{}/{}'.format(week_of, previous_folder))
            path = os.getcwd()
            for file in os.listdir(path):
                if '_Map.xlsx' in file:
                    mapname = path + "\\" + file
                    maps.append(file)
        except FileNotFoundError:
            print("Automatic gel map detection failed.")
            path = manual_directory()
            for file in os.listdir(path):
                if 'Map.xlsx' in file:
                    mapname = path + "\\" + file
                    maps.append(mapname)
    else:
        path = manual_directory()
        for file in os.listdir(path):
            if 'Map.xlsx' in file:
                mapname = path + "\\" + file
                maps.append(mapname)
    return maps


def manual_directory():
    while True:
            path = input("Enter the path of your platemap file: ")
            if not os.path.isdir(path):
                print("Not a valid directory")
            else:
                break
    return path 


def open_gel(gelfile):
    wb = load_workbook(filename=gelfile, data_only=True)
    ws = wb.worksheets[0]
    return ws


def sample_disease(ws):
    discordancies = []
    diseases = []
    failures = []
    for i in range(1, 1300):
        sample_cell = 'B{}'.format(i)
        disease_cell = 'C{}'.format(i)
        genotype_cell = 'H{}'.format(i)
        sample = ws[sample_cell].value
        disease = ws[disease_cell].value
        genotype = ws[genotype_cell].value
        if disease not in diseases:
            diseases.append(disease)
        sample_and_disease = str(sample) + ': ' + str(disease)
        if genotype == 'Fail':
            failures.append(sample_and_disease)
            if sample not in failures_master_list:
                failures_master_list.append(sample)
        else:
            if sample_and_disease not in sample_disease_dict.keys():
                if genotype != None:
                    sample_disease_dict[sample_and_disease] = genotype
            else:
                if genotype != sample_disease_dict[sample_and_disease] and sample_and_disease not in discordancies:
                    discordancies.append(sample_and_disease)
                    if sample not in discordance_master_list:
                        discordance_master_list.append(sample)
    print('\nDiscordancies')
    for b in discordancies:
        print(b)
    print('\nFailed Samples:')
    for a in failures:
        print(a)


if __name__ == "__main__":
    main()